from flask import Flask, render_template, request, send_file
import pandas as pd
import os
import calendar
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.chart import BarChart, Reference
from collections import defaultdict
import csv

# ==========================================================
# FLASK SETUP
# ==========================================================
app = Flask(__name__)  

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

OUTPUT_FILE = os.path.join(UPLOAD_FOLDER, "Identix_Attendance_Report.xlsx")

# ==========================================================
# LOAD CONFIG FROM CSV  (edit employees.csv to add/remove employees)
# ==========================================================

# Resolve CONFIG_FILE relative to this script, works on all platforms/servers
_BASE_DIR   = os.path.dirname(os.path.abspath(__file__)) if "__file__" in dir() else os.getcwd()
CONFIG_FILE = os.path.join(_BASE_DIR, "employees.csv")

# Fallback defaults used if CSV is missing or unreadable
_DEFAULT_COMPANY   = "Mignited Technologies and Solutions Private Limited"
_DEFAULT_EMPLOYEES = {
    1: "Mukund Muley",
    5: "Pralav Phakatkar",
    8: "Rohit Nimse",
    9: "Samruddhi Kulkarni",
    10: "Mohini Shendge",
    11: "Shantanu Kulkarni",
}

def load_config():
    """Read company name and employee dict from employees.csv.
    Falls back to hardcoded defaults if the file is missing or malformed."""
    company   = _DEFAULT_COMPANY
    employees = dict(_DEFAULT_EMPLOYEES)   # start with defaults

    if not os.path.exists(CONFIG_FILE):
        print(f"INFO: employees.csv not found at {CONFIG_FILE}. Using defaults.")
        return company, employees

    try:
        loaded_employees = {}
        with open(CONFIG_FILE, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for row in reader:
                if not row or all(c.strip() == "" for c in row):
                    continue
                key = row[0].strip().lower()
                if key == "company_name":
                    if len(row) > 1 and row[1].strip():
                        company = row[1].strip()
                elif key == "employee_id":
                    continue   # header row
                else:
                    try:
                        emp_id   = int(row[0].strip())
                        emp_name = row[1].strip() if len(row) > 1 else "Unknown"
                        loaded_employees[emp_id] = emp_name
                    except (ValueError, IndexError):
                        continue
        if loaded_employees:
            employees = loaded_employees   # only replace defaults if CSV had valid rows
    except Exception as e:
        print(f"WARNING: Could not read employees.csv ({e}). Using defaults.")

    return company, employees

COMPANY_NAME, EMPLOYEES = load_config()

# ==========================================================
# HELPERS
# ==========================================================
def _safe_sheet_name(name, used=None):
    """Truncate to 31 chars; if collision, shorten further to fit a suffix."""
    name = name[:31]
    if used is None:
        return name
    original = name
    counter = 1
    while name in used:
        suffix = f"_{counter}"
        name = original[:31 - len(suffix)] + suffix
        counter += 1
    used.add(name)
    return name


# ==========================================================
# LOAD DATA
# ==========================================================
def load_data(input_file):
    # File is tab-separated; DateTime is one column "YYYY-MM-DD HH:MM:SS"
    df = pd.read_csv(input_file, sep="\t", engine="python", header=None)

    # Handle both formats:
    #   6-col: EmployeeID | DateTime | VerifyMode | InOut | WorkCode | Reserved
    #   7-col: EmployeeID | Date | Time | VerifyMode | InOut | WorkCode | Reserved
    if df.shape[1] == 6:
        df.columns = ["EmployeeID", "DateTime", "VerifyMode", "InOut", "WorkCode", "Reserved"]
        df["DateTime"] = pd.to_datetime(df["DateTime"].str.strip())
    else:
        df.columns = ["EmployeeID", "Date", "Time", "VerifyMode", "InOut", "WorkCode", "Reserved"]
        df["DateTime"] = pd.to_datetime(df["Date"].str.strip() + " " + df["Time"].str.strip())

    df["EmployeeID"] = df["EmployeeID"].astype(str).str.strip().astype(int)
    df["Date"]  = df["DateTime"].dt.date
    df["Month"] = df["DateTime"].dt.strftime("%b-%Y")
    df["Year"]  = df["DateTime"].dt.year
    return df.sort_values(["EmployeeID", "DateTime"])


# ==========================================================
# DAILY ATTENDANCE
# ==========================================================
def calculate_daily(df):
    daily = (
        df.groupby(["EmployeeID", "Date", "Month", "Year"])
        .agg(In_Time=("DateTime", "first"), Out_Time=("DateTime", "last"))
        .reset_index()
    )
    daily["Work_Hours"] = daily["Out_Time"] - daily["In_Time"]
    return daily


# ==========================================================
# GET MONTH DATES  –  derived directly from month label, not from data min()
# ==========================================================
def get_month_dates(month_label):
    """month_label e.g. 'Jan-2024'"""
    dt = datetime.strptime(month_label, "%b-%Y")
    year, month = dt.year, dt.month
    total_days  = calendar.monthrange(year, month)[1]
    month_dates = [datetime(year, month, day) for day in range(1, total_days + 1)]
    sundays     = [d for d in month_dates if d.weekday() == 6]
    return month_dates, sundays


# ==========================================================
# MATRIX + SUMMARY
# ==========================================================
def create_matrix_summary(daily, month_dates, sundays):
    matrix_rows  = []
    summary_rows = []

    total_days        = len(month_dates)
    total_holidays    = len(sundays)
    base_working_days = total_days - total_holidays

    for emp_id, emp_name in EMPLOYEES.items():
        row       = {"Employee Name": emp_name}
        emp_dates = set(daily[daily["EmployeeID"] == emp_id]["Date"])
        present_count = 0
        sunday_work   = 0

        for d in month_dates:
            col = d.strftime("%d-%a")          # e.g. "01-Mon" – shorter, fits columns
            if d.date() in emp_dates:
                row[col] = "P"
                present_count += 1
                if d.weekday() == 6:
                    sunday_work += 1
            else:
                row[col] = "HOLIDAY" if d.weekday() == 6 else "A"

        total_working_days = base_working_days + sunday_work
        absent             = max(total_working_days - present_count, 0)
        attendance_percent = (
            round((present_count / total_working_days) * 100, 2)
            if total_working_days else 0
        )

        summary_rows.append({
            "Employee Name":      emp_name,
            "Total Days":         total_days,
            "Holidays (Sun)":     total_holidays,
            "Working Days":       total_working_days,
            "Present":            present_count,
            "Absent":             absent,
            "Attendance %":       attendance_percent,
        })
        matrix_rows.append(row)

    return pd.DataFrame(matrix_rows), pd.DataFrame(summary_rows)


# ==========================================================
# WRITE ATTENDANCE DETAIL SHEET  (single sheet, all months, vertical)
# ==========================================================
def _write_detail_sheet(ws, all_monthly_data):
    """
    all_monthly_data: list of (month_label, month_df) tuples, chronological.
    Writes all months into ws in downward blocks with year/month banners.
    """
    from openpyxl.utils import get_column_letter

    year_fill    = PatternFill("solid", fgColor="B8860B")
    month_fill   = PatternFill("solid", fgColor="0E6655")
    emp_fill     = PatternFill("solid", fgColor="1A5276")
    header_fill  = PatternFill("solid", fgColor="2E4057")
    alt_fill     = PatternFill("solid", fgColor="EBF5FB")

    thick_border = Border(
        left=Side(style="medium"), right=Side(style="medium"),
        top=Side(style="medium"),  bottom=Side(style="medium")
    )
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    COLS        = ["Date", "Day", "In Time", "Out Time", "Work Hours"]
    num_cols    = len(COLS)
    current_row = 1
    cur_year    = None

    def merge_row(row, value, font, fill, height, border, align):
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=num_cols)
        cell           = ws.cell(row=row, column=1, value=value)
        cell.font      = font
        cell.fill      = fill
        cell.alignment = align
        ws.row_dimensions[row].height = height
        for c in range(1, num_cols + 1):
            ws.cell(row=row, column=c).border = border

    for month_label, month_df in all_monthly_data:
        year       = datetime.strptime(month_label, "%b-%Y").year
        full_month = datetime.strptime(month_label, "%b-%Y").strftime("%B  %Y")

        # Year banner (only on year change)
        if year != cur_year:
            cur_year = year
            merge_row(current_row,
                      value  = f"  \u2605  {year}  \u2605  ",
                      font   = Font(bold=True, color="FFFFFF", size=16, name="Arial"),
                      fill   = year_fill,
                      height = 28,
                      border = thick_border,
                      align  = center)
            current_row += 1

        # Month banner
        merge_row(current_row,
                  value  = f"  {full_month}  ",
                  font   = Font(bold=True, color="FFFFFF", size=12, name="Arial"),
                  fill   = month_fill,
                  height = 22,
                  border = thin_border,
                  align  = center)
        current_row += 1

        # Employee blocks
        for emp_id, emp_data in sorted(month_df.groupby("EmployeeID"), key=lambda x: x[0]):
            emp_name = EMPLOYEES.get(emp_id, "Unknown")

            merge_row(current_row,
                      value  = f"  {emp_id}  \u2014  {emp_name}  ",
                      font   = Font(bold=True, color="FFFFFF", size=11, name="Arial"),
                      fill   = emp_fill,
                      height = 20,
                      border = thin_border,
                      align  = left)
            current_row += 1

            for ci, col_name in enumerate(COLS, start=1):
                cell           = ws.cell(row=current_row, column=ci, value=col_name)
                cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Arial")
                cell.fill      = header_fill
                cell.alignment = center
                cell.border    = thin_border
            ws.row_dimensions[current_row].height = 16
            current_row += 1

            for ri, (_, r) in enumerate(emp_data.sort_values("Date").iterrows()):
                row_vals = [
                    r["Date"],
                    r["In_Time"].strftime("%A"),
                    r["In_Time"].strftime("%H:%M:%S"),
                    r["Out_Time"].strftime("%H:%M:%S"),
                    str(r["Work_Hours"]).split(".")[0].replace("0 days ", ""),
                ]
                fill = alt_fill if ri % 2 == 1 else None
                for ci, val in enumerate(row_vals, start=1):
                    cell           = ws.cell(row=current_row, column=ci, value=val)
                    cell.alignment = center
                    cell.border    = thin_border
                    if fill:
                        cell.fill = fill
                ws.row_dimensions[current_row].height = 15
                current_row += 1

            current_row += 1   # spacer between employees

        current_row += 2       # spacer between months

    # Auto column widths
    for ci in range(1, num_cols + 1):
        col_letter = get_column_letter(ci)
        max_len = max(
            (len(str(ws.cell(row=r, column=ci).value))
             if ws.cell(row=r, column=ci).value else 0
             for r in range(1, ws.max_row + 1)),
            default=0
        )
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    ws.freeze_panes = None



# ==========================================================
# WRITE EXCEL
# ==========================================================
def write_excel(daily, all_matrix_data, final_summary, month_order):
    """
    all_matrix_data : list of (month_label, matrix_df) – chronological
    final_summary   : combined df with 'Month' column
    month_order     : sorted list of month labels (same order as all_matrix_data)
    """
    if os.path.exists(OUTPUT_FILE):
        try:
            os.remove(OUTPUT_FILE)
        except PermissionError:
            print("Please close the Excel file and try again.")
            return

    used_names = set()

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

        # Single Attendance Detail sheet (all months, all employees, vertical)
        ws_det = writer.book.create_sheet("Attendance_Detail")
        used_names.add("Attendance_Detail")
        all_monthly = [(m, daily[daily["Month"] == m]) for m in month_order]
        _write_detail_sheet(ws_det, all_monthly)


        # ── Single Attendance Matrix sheet (placeholder – filled below) ──────
        ws_matrix = writer.book.create_sheet("Attendance_Matrix")
        ws_matrix.append(["placeholder"])
        used_names.add("Attendance_Matrix")

        # Single Attendance_Summary sheet placeholder (filled after ExcelWriter closes)
        ws_sum = writer.book.create_sheet("Attendance_Summary")
        ws_sum.append(["placeholder"])
        used_names.add("Attendance_Summary")

    # ── Populate Attendance_Matrix via openpyxl ──────────────────────────────
    _write_matrix_sheet(all_matrix_data)
    _write_summary_sheet(final_summary, month_order)

    # Enforce tab order: Summary → Detail → Matrix
    wb = load_workbook(OUTPUT_FILE)
    desired = ["Attendance_Summary", "Attendance_Detail", "Attendance_Matrix"]
    wb._sheets.sort(key=lambda s: desired.index(s.title) if s.title in desired else len(desired))
    wb.save(OUTPUT_FILE)



# ==========================================================
# WRITE ATTENDANCE MATRIX SHEET
# ==========================================================
def _write_matrix_sheet(all_matrix_data):
    """Writes all months (all years) into a single Attendance_Matrix sheet."""
    wb = load_workbook(OUTPUT_FILE)
    ws = wb["Attendance_Matrix"]
    ws.delete_rows(1, ws.max_row)
    year_fill    = PatternFill("solid", fgColor="B8860B")   # dark gold for year
    month_fill   = PatternFill("solid", fgColor="0E6655")   # deep teal for month
    header_fill  = PatternFill("solid", fgColor="1A5276")   # steel blue for col headers
    green_fill   = PatternFill("solid", fgColor="C6EFCE")
    red_fill     = PatternFill("solid", fgColor="FFC7CE")
    grey_fill    = PatternFill("solid", fgColor="D9D9D9")
    border       = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    year_border  = Border(
        left=Side(style='medium'), right=Side(style='medium'),
        top=Side(style='medium'),  bottom=Side(style='medium')
    )
    month_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='medium')
    )
    center = Alignment(horizontal="center", vertical="center")

    current_row  = 1
    current_year = None

    for month_label, matrix_df in all_matrix_data:
        year     = datetime.strptime(month_label, "%b-%Y").year
        num_cols = len(matrix_df.columns)

        # ── Year banner (only when year changes) ─────────────────────────
        if year != current_year:
            current_year = year
            cell = ws.cell(row=current_row, column=1, value=f"  ★  {year}  ★  ")
            cell.font      = Font(bold=True, color="FFFFFF", size=16, name="Arial")
            cell.fill      = year_fill
            cell.alignment = center
            ws.row_dimensions[current_row].height = 28
            if num_cols > 1:
                ws.merge_cells(
                    start_row=current_row, start_column=1,
                    end_row=current_row,   end_column=num_cols
                )
            for c in range(1, num_cols + 1):
                ws.cell(row=current_row, column=c).border = year_border
            current_row += 1

        # ── Month title row ───────────────────────────────────────────────
        full_month = datetime.strptime(month_label, "%b-%Y").strftime("%B  %Y")
        cell = ws.cell(row=current_row, column=1, value=f"  {full_month}  ")
        cell.font      = Font(bold=True, color="FFFFFF", size=12, name="Arial")
        cell.fill      = month_fill
        cell.alignment = center
        ws.row_dimensions[current_row].height = 22
        if num_cols > 1:
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row,   end_column=num_cols
            )
        for c in range(1, num_cols + 1):
            ws.cell(row=current_row, column=c).border = month_border
        current_row += 1


        # ── Column header row ─────────────────────────────────────────────
        for col_idx, col_name in enumerate(matrix_df.columns, start=1):
            cell            = ws.cell(row=current_row, column=col_idx, value=col_name)
            cell.fill       = header_fill
            cell.font       = Font(bold=True, color="FFFFFF")
            cell.alignment  = center
            cell.border     = border
        current_row += 1

        # ── Data rows ─────────────────────────────────────────────────────
        for _, data_row in matrix_df.iterrows():
            for col_idx, value in enumerate(data_row, start=1):
                cell           = ws.cell(row=current_row, column=col_idx, value=value)
                cell.alignment = center
                cell.border    = border
                if col_idx > 1:
                    if value == "P":
                        cell.fill = green_fill
                    elif value == "A":
                        cell.fill = red_fill
                    elif value == "HOLIDAY":
                        cell.fill = grey_fill
            current_row += 1

        # ── Spacer between months ─────────────────────────────────────────
        current_row += 2

    # ── Auto column widths ────────────────────────────────────────────────
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value))
             if ws.cell(row=r, column=col_idx).value else 0
             for r in range(1, ws.max_row + 1)),
            default=0
        )
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    ws.freeze_panes = "B1"
    wb.save(OUTPUT_FILE)


# ==========================================================
# FORMAT EXCEL  (detail + summary sheets only)
# ==========================================================
def format_excel():
    wb = load_workbook(OUTPUT_FILE)

    header_fill = PatternFill("solid", fgColor="305496")
    border      = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    # Skip sheets that are already self-formatted
    month_detail_sheets = set()
    for m in [f"{d.strftime('%b')}-{d.year}" for d in 
              [datetime(y, mo, 1) for y in range(2000, 2100) for mo in range(1, 13)]]:
        month_detail_sheets.add(m)
    skip = {"Attendance_Matrix"}
    # Add all month-detail sheet names (format: "Jan-2024") to skip
    import re
    month_re = re.compile(r'^[A-Z][a-z]{2}-\d{4}$')
    skip = {"Attendance_Matrix", "Attendance_Detail", "Attendance_Summary"}

    for sheet in wb.sheetnames:
        if sheet in skip or sheet.startswith("Dashboard_"):
            continue

        ws = wb[sheet]
        ws.freeze_panes = "A2"

        for row in ws.iter_rows():
            for cell in row:
                cell.border    = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")

        for col in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value else 0
                for cell in col
            )
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)

    wb.save(OUTPUT_FILE)


# ==========================================================
# DASHBOARD  –  one per month, uses summary_sheet_map for safe names
# ==========================================================
# WRITE SUMMARY SHEET  (single sheet, all months, vertical)
# ==========================================================
# ==========================================================
# WRITE SUMMARY SHEET  (with KPI cards + bar chart per month)
# ==========================================================
def _write_summary_sheet(final_summary, month_order):
    from openpyxl.utils import get_column_letter
    wb = load_workbook(OUTPUT_FILE)
    ws = wb["Attendance_Summary"]
    ws.delete_rows(1, ws.max_row)

    year_fill    = PatternFill("solid", fgColor="B8860B")
    month_fill   = PatternFill("solid", fgColor="0E6655")
    header_fill  = PatternFill("solid", fgColor="2E4057")
    good_fill    = PatternFill("solid", fgColor="C6EFCE")
    bad_fill     = PatternFill("solid", fgColor="FFC7CE")
    alt_fill     = PatternFill("solid", fgColor="EBF5FB")
    thick_bdr    = Border(left=Side(style="medium"), right=Side(style="medium"),
                          top=Side(style="medium"),  bottom=Side(style="medium"))
    thin_bdr     = Border(left=Side(style="thin"),   right=Side(style="thin"),
                          top=Side(style="thin"),    bottom=Side(style="thin"))
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    COLS     = ["Employee Name", "Total Days", "Holidays (Sun)", "Working Days",
                "Present", "Absent", "Attendance %"]
    NUM_COLS = len(COLS)
    cur_row  = 1
    cur_year = None

    # Company header
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=NUM_COLS)
    c = ws.cell(row=cur_row, column=1, value=f"  {COMPANY_NAME}  —  Attendance Summary  ")
    c.font = Font(bold=True, color="FFFFFF", size=14, name="Arial")
    c.fill = PatternFill("solid", fgColor="1B2631")
    c.alignment = center
    ws.row_dimensions[cur_row].height = 30
    for col in range(1, NUM_COLS + 1):
        ws.cell(row=cur_row, column=col).border = thick_bdr
    cur_row += 2   # spacer after company header

    for month_label in month_order:
        year       = datetime.strptime(month_label, "%b-%Y").year
        full_month = datetime.strptime(month_label, "%b-%Y").strftime("%B  %Y")
        mdf        = final_summary[final_summary["Month"] == month_label]
        pcts       = list(mdf["Attendance %"])
        names      = list(mdf["Employee Name"])
        best_pct   = max(pcts) if pcts else None
        worst_pct  = min(pcts) if pcts else None

        # Year banner (on year change)
        if year != cur_year:
            cur_year = year
            ws.merge_cells(start_row=cur_row, start_column=1,
                           end_row=cur_row,   end_column=NUM_COLS)
            c = ws.cell(row=cur_row, column=1, value=f"  ★  {year}  ★  ")
            c.font = Font(bold=True, color="FFFFFF", size=16, name="Arial")
            c.fill = year_fill; c.alignment = center
            ws.row_dimensions[cur_row].height = 28
            for col in range(1, NUM_COLS + 1):
                ws.cell(row=cur_row, column=col).border = thick_bdr
            cur_row += 1

        # Month banner
        ws.merge_cells(start_row=cur_row, start_column=1,
                       end_row=cur_row,   end_column=NUM_COLS)
        c = ws.cell(row=cur_row, column=1, value=f"  {full_month}  ")
        c.font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
        c.fill = month_fill; c.alignment = center
        ws.row_dimensions[cur_row].height = 22
        for col in range(1, NUM_COLS + 1):
            ws.cell(row=cur_row, column=col).border = thin_bdr
        cur_row += 1

        # Column headers
        kpi_start_row = cur_row   # chart anchors here (top of data block)
        for ci, col_name in enumerate(COLS, start=1):
            c = ws.cell(row=cur_row, column=ci, value=col_name)
            c.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
            c.fill = header_fill; c.alignment = center; c.border = thin_bdr
        ws.row_dimensions[cur_row].height = 16
        cur_row += 1

        # Data rows
        data_start = cur_row
        for ri, (_, row) in enumerate(mdf.iterrows()):
            vals = [row["Employee Name"], row["Total Days"], row["Holidays (Sun)"],
                    row["Working Days"], row["Present"], row["Absent"], row["Attendance %"]]
            base_fill = alt_fill if ri % 2 == 1 else None
            for ci, val in enumerate(vals, start=1):
                c = ws.cell(row=cur_row, column=ci, value=val)
                c.alignment = center; c.border = thin_bdr
                if ci == 7:
                    c.fill = good_fill if val == best_pct else (bad_fill if val == worst_pct else (base_fill or PatternFill()))
                else:
                    if base_fill: c.fill = base_fill
            ws.row_dimensions[cur_row].height = 15
            cur_row += 1

        data_end = cur_row - 1

        # Bar chart — anchored to column I at KPI start row
        chart              = BarChart()
        chart.title        = f"Attendance % — {full_month}"
        chart.y_axis.title = "Attendance %"
        chart.width        = 18
        chart.height       = 10
        chart.add_data(Reference(ws, min_col=7, min_row=data_start, max_row=data_end))
        chart.set_categories(Reference(ws, min_col=1, min_row=data_start, max_row=data_end))
        ws.add_chart(chart, f"I{kpi_start_row}")

        cur_row += 2   # spacer between months

    # Auto widths
    for ci in range(1, NUM_COLS + 1):
        col_letter = get_column_letter(ci)
        max_len = max(
            (len(str(ws.cell(row=r, column=ci).value))
             if ws.cell(row=r, column=ci).value else 0
             for r in range(1, ws.max_row + 1)), default=0)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 16)

    ws.freeze_panes = None
    wb.save(OUTPUT_FILE)



# REPORT GENERATOR
# ==========================================================
def generate_report(input_file):
    df    = load_data(input_file)
    daily = calculate_daily(df)

    # Chronological sort across multiple years: sort by actual date
    month_order = sorted(
        daily["Month"].unique(),
        key=lambda m: datetime.strptime(m, "%b-%Y"),
        reverse=True
    )
    all_matrix_data = []
    all_summary     = []

    for month in month_order:
        month_df              = daily[daily["Month"] == month]
        month_dates, sundays  = get_month_dates(month)
        matrix, summary       = create_matrix_summary(month_df, month_dates, sundays)
        summary.insert(0, "Month", month)
        all_matrix_data.append((month, matrix))
        all_summary.append(summary)

    final_summary    = pd.concat(all_summary, ignore_index=True)
    write_excel(daily, all_matrix_data, final_summary, month_order)
    format_excel()

    return OUTPUT_FILE


# ==========================================================
# CSV HELPERS
# ==========================================================
def save_config(company, employees):
    """Write company name and employees back to employees.csv."""
    with open(CONFIG_FILE, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["company_name", company])
        writer.writerow(["employee_id", "employee_name"])
        for emp_id, emp_name in sorted(employees.items()):
            writer.writerow([emp_id, emp_name])


# ==========================================================
# FLASK ROUTES
# ==========================================================
@app.route("/")
def home():
    return render_template("index.html")


@app.route("/employees")
def employees_page():
    global COMPANY_NAME, EMPLOYEES
    COMPANY_NAME, EMPLOYEES = load_config()
    rows = [{"id": k, "name": v} for k, v in sorted(EMPLOYEES.items())]
    return render_template("employees.html",
                           company=COMPANY_NAME,
                           employees=rows)


@app.route("/employees/add", methods=["POST"])
def add_employee():
    global COMPANY_NAME, EMPLOYEES
    COMPANY_NAME, EMPLOYEES = load_config()
    try:
        emp_id   = int(request.form["emp_id"].strip())
        emp_name = request.form["emp_name"].strip()
        if not emp_name:
            return "Employee name cannot be empty", 400
        EMPLOYEES[emp_id] = emp_name
        save_config(COMPANY_NAME, EMPLOYEES)
    except (ValueError, KeyError):
        return "Invalid employee ID", 400
    return ("", 204)


@app.route("/employees/edit", methods=["POST"])
def edit_employee():
    global COMPANY_NAME, EMPLOYEES
    COMPANY_NAME, EMPLOYEES = load_config()
    try:
        emp_id   = int(request.form["emp_id"].strip())
        emp_name = request.form["emp_name"].strip()
        if emp_id not in EMPLOYEES:
            return "Employee not found", 404
        if not emp_name:
            return "Employee name cannot be empty", 400
        EMPLOYEES[emp_id] = emp_name
        save_config(COMPANY_NAME, EMPLOYEES)
    except (ValueError, KeyError):
        return "Invalid input", 400
    return ("", 204)


@app.route("/employees/delete", methods=["POST"])
def delete_employee():
    global COMPANY_NAME, EMPLOYEES
    COMPANY_NAME, EMPLOYEES = load_config()
    try:
        emp_id = int(request.form["emp_id"].strip())
        EMPLOYEES.pop(emp_id, None)
        save_config(COMPANY_NAME, EMPLOYEES)
    except (ValueError, KeyError):
        return "Invalid employee ID", 400
    return ("", 204)


@app.route("/employees/company", methods=["POST"])
def update_company():
    global COMPANY_NAME, EMPLOYEES
    COMPANY_NAME, EMPLOYEES = load_config()
    name = request.form.get("company_name", "").strip()
    if not name:
        return "Company name cannot be empty", 400
    COMPANY_NAME = name
    save_config(COMPANY_NAME, EMPLOYEES)
    return ("", 204)


@app.route("/upload", methods=["POST"])
def upload_file():
    try:
        file = request.files.get("file")
        if not file or file.filename == "":
            return "No file uploaded", 400

        global COMPANY_NAME, EMPLOYEES
        COMPANY_NAME, EMPLOYEES = load_config()

        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)

        # Merge uploaded file into master log (deduped)
        added = merge_into_master_log(filepath)
        print(f"Merged {added} new punches into master log")

        # Always generate report from full master log
        generate_report(LIVE_LOG)
        return render_template("dashboard.html", filename="Identix_Attendance_Report.xlsx")

    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"<h2>Error generating report</h2><pre>{traceback.format_exc()}</pre>", 500


@app.route("/download/<filename>")
def download_file(filename):
    return send_file(os.path.join(UPLOAD_FOLDER, filename), as_attachment=True)


# ==========================================================
# DATE FILTER HELPER
# ==========================================================
def _passes_filter(dt, filter_type, now, custom_start=None, custom_end=None):
    """Return True if datetime dt passes the given filter."""
    from datetime import timedelta
    if filter_type == "today":
        return dt.date() == now.date()
    elif filter_type == "week":
        return (now.date() - dt.date()).days <= 6
    elif filter_type == "last_week":
        start = now.date() - timedelta(days=now.weekday() + 7)
        end   = start + timedelta(days=6)
        return start <= dt.date() <= end
    elif filter_type == "month":
        return dt.month == now.month and dt.year == now.year
    elif filter_type == "last_month":
        if now.month == 1:
            return dt.month == 12 and dt.year == now.year - 1
        return dt.month == now.month - 1 and dt.year == now.year
    elif filter_type == "year":
        return dt.year == now.year
    elif filter_type == "custom":
        if custom_start and custom_end:
            return custom_start <= dt.date() <= custom_end
        return True
    return True   # "all"


# ==========================================================
# LIVE BIOMETRIC RECEIVER  (Identix K21 / K30 Pro ADMS push)
# ==========================================================
LIVE_LOG = os.path.join(UPLOAD_FOLDER, "live_attendance.dat")
_regen_lock = False   # prevent overlapping report regenerations


def _load_master_log():
    """Load all existing punch lines from master log as a set of tuples for dedup."""
    existing = set()
    if os.path.exists(LIVE_LOG):
        with open(LIVE_LOG, encoding="utf-8") as f:
            for line in f:
                parts = line.strip().split("\t")
                if len(parts) >= 2:
                    existing.add((parts[0].strip(), parts[1].strip()))
    return existing


def merge_into_master_log(dat_file):
    """Merge all punches from a .dat file into live_attendance.dat, no duplicates."""
    existing = _load_master_log()
    new_lines = []
    try:
        with open(dat_file, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                parts = line.split("\t")
                if len(parts) < 2:
                    # try space-separated
                    parts = line.split()
                if len(parts) < 2:
                    continue
                emp_id   = parts[0].strip()
                datetime_str = parts[1].strip()
                # For space-sep format: date and time are separate tokens
                if len(datetime_str) == 10 and len(parts) >= 3:
                    datetime_str = f"{datetime_str} {parts[2].strip()}"
                key = (emp_id, datetime_str)
                if key not in existing:
                    existing.add(key)
                    # Rebuild as tab-separated with 6 columns
                    verify   = parts[3].strip() if len(parts) > 3 else "1"
                    inout    = parts[4].strip() if len(parts) > 4 else "0"
                    workcode = parts[5].strip() if len(parts) > 5 else "0"
                    new_lines.append(
                        f"{emp_id}\t{datetime_str}\t{verify}\t{inout}\t{workcode}\t0\n"
                    )
    except Exception as e:
        print(f"merge_into_master_log error: {e}")
        return 0

    if new_lines:
        with open(LIVE_LOG, "a", encoding="utf-8") as f:
            f.writelines(new_lines)

    return len(new_lines)


def _append_punch(emp_id, datetime_str, verify, inout, workcode):
    """Append a single live punch to master log if not duplicate."""
    existing = _load_master_log()
    key = (str(emp_id), datetime_str)
    if key not in existing:
        with open(LIVE_LOG, "a", encoding="utf-8") as f:
            f.write(f"{emp_id}\t{datetime_str}\t{verify}\t{inout}\t{workcode}\t0\n")


def _safe_regen():
    """Regenerate the Excel report from live_attendance.dat safely."""
    global _regen_lock
    if _regen_lock:
        return
    _regen_lock = True
    try:
        global COMPANY_NAME, EMPLOYEES
        COMPANY_NAME, EMPLOYEES = load_config()
        if os.path.exists(LIVE_LOG) and os.path.getsize(LIVE_LOG) > 0:
            generate_report(LIVE_LOG)   # always from master log
    except Exception as e:
        import traceback
        traceback.print_exc()
    finally:
        _regen_lock = False


@app.route("/iclock/cdata", methods=["GET", "POST"])
def biometric_receiver():
    """
    Identix ADMS endpoint. Device sends punches here automatically.
    GET  — device handshake / info request
    POST — actual attendance punch data
    """
    sn = request.args.get("SN") or request.form.get("SN", "DEVICE")

    if request.method == "GET":
        # Respond with server time so device stays in sync
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return (f"GET SUCCESSFUL\nStamp=9999\nServerTime={now}\n"), 200

    # POST — parse punch records
    # Device sends: SN=xxx&table=ATTLOG&Stamp=1&Data=EmpID Date Time Verify InOut WorkCode
    table = request.form.get("table", "")
    data  = request.form.get("Data", "")

    if table == "ATTLOG" and data:
        punches_added = 0
        for line in data.strip().splitlines():
            parts = line.strip().split()
            if len(parts) >= 5:
                try:
                    emp_id      = parts[0].strip()
                    date_str    = parts[1].strip()
                    time_str    = parts[2].strip()
                    verify      = parts[3].strip()
                    inout       = parts[4].strip()
                    workcode    = parts[5].strip() if len(parts) > 5 else "0"
                    datetime_str = f"{date_str} {time_str}"
                    _append_punch(emp_id, datetime_str, verify, inout, workcode)
                    punches_added += 1
                except Exception:
                    continue

        if punches_added > 0:
            # Regenerate report in background thread so device doesn't timeout
            import threading
            threading.Thread(target=_safe_regen, daemon=True).start()

    return "OK", 200


@app.route("/live")
def live_status():
    """Shows live punch log with filter options."""
    from datetime import timedelta
    filter_type  = request.args.get("filter", "today")
    custom_start = None
    custom_end   = None
    now          = datetime.now()

    # Parse custom date range
    if filter_type == "custom":
        try:
            from datetime import date as date_cls
            custom_start = date_cls.fromisoformat(request.args.get("start", ""))
            custom_end   = date_cls.fromisoformat(request.args.get("end", ""))
        except ValueError:
            filter_type = "all"

    punches = []
    if os.path.exists(LIVE_LOG):
        with open(LIVE_LOG, encoding="utf-8") as f:
            for line in f.readlines():
                parts = line.strip().split("\t")
                if len(parts) < 2:
                    continue
                try:
                    dt = datetime.strptime(parts[1].strip(), "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    continue
                if not _passes_filter(dt, filter_type, now, custom_start, custom_end):
                    continue
                punches.append({
                    "emp_id":   parts[0].strip(),
                    "datetime": parts[1].strip(),
                    "name":     EMPLOYEES.get(int(parts[0].strip()), "Unknown")
                              if parts[0].strip().isdigit() else "Unknown"
                })

    punches.reverse()
    return render_template("live.html",
                           punches=punches,
                           filter_type=filter_type,
                           custom_start=str(custom_start) if custom_start else "",
                           custom_end=str(custom_end)   if custom_end   else "",
                           report_exists=os.path.exists(OUTPUT_FILE))


@app.route("/live/download")
def live_download():
    """Generate and download Excel report filtered by date range."""
    import tempfile
    filter_type = request.args.get("filter", "today")
    now         = datetime.now()

    if not os.path.exists(LIVE_LOG) or os.path.getsize(LIVE_LOG) == 0:
        return "No live data available yet.", 400

    # Parse custom dates if needed
    custom_start = None
    custom_end   = None
    if filter_type == "custom":
        try:
            from datetime import date as date_cls
            custom_start = date_cls.fromisoformat(request.args.get("start", ""))
            custom_end   = date_cls.fromisoformat(request.args.get("end", ""))
        except ValueError:
            filter_type = "all"

    # Filter lines from live log into a temp file
    filtered_lines = []
    with open(LIVE_LOG, encoding="utf-8") as f:
        for line in f.readlines():
            parts = line.strip().split("\t")
            if len(parts) < 2:
                continue
            try:
                dt = datetime.strptime(parts[1].strip(), "%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
            if not _passes_filter(dt, filter_type, now, custom_start, custom_end):
                continue
            filtered_lines.append(line)

    if not filtered_lines:
        return "No data found for the selected period.", 400

    # Write filtered lines to temp file and generate report
    tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".dat",
                                      delete=False, encoding="utf-8")
    tmp.writelines(filtered_lines)
    tmp.close()

    try:
        global COMPANY_NAME, EMPLOYEES, OUTPUT_FILE
        COMPANY_NAME, EMPLOYEES = load_config()

        filter_labels = {
            "today":     "Today",
            "week":      "This_Week",
            "last_week": "Last_Week",
            "month":     "This_Month",
            "last_month":"Last_Month",
            "year":      "This_Year",
            "custom":    f"Custom_{request.args.get('start','')}_{request.args.get('end','')}",
            "all":       "All_Data"
        }
        label    = filter_labels.get(filter_type, filter_type)
        out_file = os.path.join(UPLOAD_FOLDER,
                                f"Attendance_Report_{label}.xlsx")

        # Temporarily swap OUTPUT_FILE so generate_report writes to out_file
        orig_output = OUTPUT_FILE
        OUTPUT_FILE = out_file

        try:
            generate_report(tmp.name)
        finally:
            OUTPUT_FILE = orig_output   # always restore

        if not os.path.exists(out_file):
            return "Report could not be generated. Please upload data first.", 400

        return send_file(out_file, as_attachment=True,
                         download_name=f"Attendance_Report_{label}.xlsx")
    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"<pre>{traceback.format_exc()}</pre>", 500
    finally:
        try:
            os.unlink(tmp.name)
        except Exception:
            pass



# ==========================================================
# PULL DATA FROM DEVICE  (ZKTeco/Identix TCP protocol)
# ==========================================================

DEVICE_IP       = "192.168.0.126"
DEVICE_PORT     = 4370
DEVICE_PASSWORD = 0   # default; change if device has different password


def _pull_zk_data(filter_type="all", **kwargs):
    """
    Connect to Identix device via pyzk, pull attendance logs filtered
    by date range, merge into master log.
    Returns (added, total_on_device, filtered_count, error_msg).
    """
    try:
        from zk import ZK
    except ImportError:
        return 0, 0, 0, "pyzk not installed. Run: pip install pyzk"

    zk = ZK(DEVICE_IP, port=DEVICE_PORT, timeout=10,
             password=DEVICE_PASSWORD, force_udp=False, ommit_ping=False)
    conn = None
    now  = datetime.now()

    try:
        conn = zk.connect()
        conn.disable_device()

        attendances  = conn.get_attendance()
        total_device = len(attendances)

        if total_device == 0:
            return 0, 0, 0, "No attendance records found on device."

        # Parse custom dates
        custom_start = None
        custom_end   = None
        if filter_type == "custom":
            try:
                from datetime import date as date_cls
                cs = kwargs.get("custom_start")
                ce = kwargs.get("custom_end")
                custom_start = date_cls.fromisoformat(cs) if cs else None
                custom_end   = date_cls.fromisoformat(ce) if ce else None
            except (ValueError, AttributeError):
                filter_type = "all"

        # Apply date filter
        filtered = []
        for att in attendances:
            dt = att.timestamp
            if _passes_filter(dt, filter_type, now, custom_start, custom_end):
                filtered.append(att)

        filtered_count = len(filtered)
        if filtered_count == 0:
            return 0, total_device, 0, None

        import tempfile
        tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".dat",
                                          delete=False, encoding="utf-8")
        for att in filtered:
            dt_str = att.timestamp.strftime("%Y-%m-%d %H:%M:%S")
            tmp.write(f"{att.user_id}\t{dt_str}\t1\t{att.punch}\t0\t0\n")
        tmp.close()

        added = merge_into_master_log(tmp.name)
        os.unlink(tmp.name)

        return added, total_device, filtered_count, None

    except Exception as e:
        return 0, 0, 0, str(e)
    finally:
        if conn:
            try:
                conn.enable_device()
                conn.disconnect()
            except Exception:
                pass


@app.route("/pull-device", methods=["GET"])
def pull_device_page():
    """Show filter selection page before pulling data."""
    return render_template("pull_device.html",
                           device_ip=DEVICE_IP,
                           device_port=DEVICE_PORT)


@app.route("/pull-device", methods=["POST"])
def pull_device():
    """Pull attendance data from biometric device filtered by date range."""
    try:
        global COMPANY_NAME, EMPLOYEES
        COMPANY_NAME, EMPLOYEES = load_config()

        filter_type = request.form.get("filter", "all")
        filter_labels = {
            "today":     "Today",
            "week":      "This Week",
            "last_week": "Last Week",
            "month":     "This Month",
            "last_month":"Last Month",
            "year":      "This Year",
            "custom":    "Custom Range",
            "all":       "All Data"
        }
        filter_label = filter_labels.get(filter_type, "All Data")

        custom_start = request.form.get("start", "")
        custom_end   = request.form.get("end", "")
        added, total_device, filtered_count, error = _pull_zk_data(
            filter_type, custom_start=custom_start, custom_end=custom_end
        )

        if error:
            return render_template("pull_device.html",
                                   device_ip=DEVICE_IP,
                                   device_port=DEVICE_PORT,
                                   error=error), 500

        # Regenerate Excel from updated master log
        if added > 0 and os.path.exists(LIVE_LOG):
            generate_report(LIVE_LOG)

        return render_template("pull_device.html",
                               device_ip=DEVICE_IP,
                               device_port=DEVICE_PORT,
                               success=True,
                               filter_label=filter_label,
                               filter_type=filter_type,
                               total_device=total_device,
                               filtered_count=filtered_count,
                               added=added,
                               skipped=filtered_count - added)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"<h2>Error</h2><pre>{traceback.format_exc()}</pre>", 500


@app.route("/device-settings", methods=["GET", "POST"])
def device_settings():
    """Page to update device IP, port and password."""
    global DEVICE_IP, DEVICE_PORT, DEVICE_PASSWORD
    message = ""

    if request.method == "POST":
        try:
            DEVICE_IP       = request.form.get("device_ip", DEVICE_IP).strip()
            DEVICE_PORT     = int(request.form.get("device_port", DEVICE_PORT))
            DEVICE_PASSWORD = int(request.form.get("device_password", DEVICE_PASSWORD))
            message = "✅ Settings saved!"
        except ValueError:
            message = "❌ Invalid port or password — must be numbers."

    return render_template("device_settings.html",
                           device_ip=DEVICE_IP,
                           device_port=DEVICE_PORT,
                           device_password=DEVICE_PASSWORD,
                           message=message)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)